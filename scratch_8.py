from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
import openpyxl
import numpy as np
import requests
import time
from bs4 import BeautifulSoup
path='/Users/yashjain/Downloads/chromedriver'
driver=webdriver.Chrome(path)
faculty_list=driver.get("https://telfer.uottawa.ca/en/directory/professors/")
url="https://telfer.uottawa.ca/en/directory/professors/"
source=requests.get(url).text
faculty_soup=BeautifulSoup(source,'lxml')
links=faculty_soup.findAll('div',class_='col-md-4 col-5')
y=0
faculty_name_list=[]
website=[]
email=[]
cv=[]
position=[]
dept=[]
name=[]
expertise=[]

#wb=openpyxl.Workbook()
#ws1 = wb.create_sheet("London School of Economics ")
#ws1.title = "Schulich"

for link in links:
    faculty_list = link.find('a')
    try:
        x=str(faculty_list.attrs['title'])
        print x[9:]
        faculty_name_list.append(x[9:])
    except:
        pass

for i in range (0,len(faculty_name_list)):

    ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
    y+=1
    try:
        q='//*[@id="directory-container"]/div[4]/div['+str(i+1)+']/div/div/div/div[2]/p/a'

        mai = WebDriverWait(driver, 5,ignored_exceptions=ignored_exceptions).until(
            EC.presence_of_element_located((By.XPATH, q)))
        print mai
        mai.click()
        print 1
        urlpage = driver.current_url
        #print urlpage
        source = requests.get(urlpage).text
        soup = BeautifulSoup(source, 'lxml')
        print(soup.findAll('div', class_='col-auto'))
        expertise = soup.find(id='research_areas')
        print expertise.text





        basic_info = soup.find('div', class_='col')
        position.append(basic_info.find('div',class_='prof_title_2').text)
        print 22
        contact_details = soup.find('a', class_="col-auto")
        website.append(contact_details.attrs['href'])
        research=(soup.find(id='research_areas'))
        areas=research.findAll('li',class_="list-group-item")
        for area in areas:
            expertise.append(area)
        print position
        #print contact_details
        print website
        print expertise





        name.append(basic_info.h1.text)

        dept.append(basic_info.h3.text)




        desc = soup.find('div', class_='people__bio')
        cv.append((desc.find('a', href=True)).attrs['href'])
        expertise.append((soup.find('div', class_='expertise')).text)


        driver.back()
        y=y+1
    except:
        driver.back()
        pass
counter=0
for i in range(1,len(faculty_name_list)-1):
    counter+=1
    try:
        c1=ws1.cell(row=i,column=1)
        c1.value=name[i]
    except:
        print i

    try:
        c2 = ws1.cell(row=i, column=2)
        c2.value = position[i]
    except:
        print i
        print name[i]
    try:
        c3 = ws1.cell(row=i, column=3)
        c3.value = dept[i]
    except:
        print i
        print name[i]
    try:
        c5 = ws1.cell(row=i, column=5)
        c5.value = website[i]
    except:
        print i
        print name[i]
    try:
        c6 = ws1.cell(row=i, column=6)
        c6.value = email[i]
    except:
        print i
        print name[i]
    try:
        c7 = ws1.cell(row=i, column=7)
        c7.value = cv[i]
    except:
        print i
        print name[i]
    try:
        c8 = ws1.cell(row=i, column=7)
        c8.value = expertise[i]
    except:
        print name[i]
    print counter
#wb.template=False
#wb.save(filename="Forgien.xlsx")
print name
print position
print dept
print website
print email
print cv
print expertise







#soup=BeautifulSoup(x,'lmxl')
#x=soup.find('div',class_='siteSearch__container')
#print x
#print (driver.title)
#time.sleep(5)
#driver.quit()
#//*[@id="directory-container"]/div[4]/div[1]/div/div/div/div[2]/p/a
#//*[@id="directory-container"]/div[4]/div[2]/div/div/div/div[2]/p/a
#/html/body/div[1]/div[3]/div[4]/div[2]/div/div/div/div[2]/p/a
#/html/body/div[1]/div[3]/div[4]/div[1]/div/div/div/div[2]/p/a