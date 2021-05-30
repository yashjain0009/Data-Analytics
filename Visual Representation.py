import openpyxl
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import scipy.stats as stats
from scipy.stats import spearmanr
path="/Users/yashjain/Desktop/Pharma_Dataset (1).xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
Return=[]
RD=[]
for i in range(2,62):
    Return.append(((sheet_obj.cell(row = i, column = 4).value)))
    RD.append(((sheet_obj.cell(row = i, column = 6).value)))
print RD
Return = np.array(Return)
RD=np.array(RD)
mu=Return.mean()
std= Return.std()
x=np.linspace(mu-3*std,mu+3*std,100)
plt.plot(x,stats.norm.pdf(x,mu,std))
y=np.linspace(-1.0,2,10)
x=(y*0)+mu
plt.plot(x,y,color='red')
x2=np.linspace(-1.5,2,10)
y2=RD.mean()
y2=(x2*0)+y2
plt.plot(x2,y2,color='red')
plt.scatter(Return,RD,marker="*")
plt.xlabel("Annual Return")
plt.ylabel("Growth in R&D Expenditure Prev-Year")
plt.tight_layout()
plt.grid(True)
plt.legend(loc=1)
plt.show()
#print np.corrcoef(Return, RD)



