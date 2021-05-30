import openpyxl
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import scipy.stats as stats
from scipy.stats import spearmanr
from scipy.stats import pearsonr
from scipy.stats import skew
from scipy.stats import kurtosis
path="/Users/yashjain/Downloads/Pharma Research /Pharma_2009-2019_Dataset_Pooled.xlsx"
wb_obj = openpyxl.load_workbook(path)
#wb_obj.template=True
sheet_obj = wb_obj.active
sheets = wb_obj.sheetnames
ws = wb_obj[sheets[2]]
#pooled_dataset=wb_obj["Sheet1"]
Company=[]
date=[]
RD=[]
current_assets=[]
current_liabilities=[]
ROA=[]
ROE=[]
ROC=[]
CR=[]
QR=[]
DER=[]
CP=[]
MCap=[]
Total_Returns=[]
EPS=[]
date2=[]
x=2
y=0
z=0
x1=5
while z<=6:
    date.append(sheet_obj.cell(row=x,column=2).value)
    Company.append((sheet_obj.cell(row=x,column=1).value))
    RD.append(sheet_obj.cell(row=x,column=3).value)
    current_assets.append(sheet_obj.cell(row=x1,column=4).value)
    current_liabilities.append(sheet_obj.cell(row=x1,column=5).value)
    ROA.append(sheet_obj.cell(row=x1,column=6).value)
    ROE.append(sheet_obj.cell(row=x1,column=7).value)
    ROC.append(sheet_obj.cell(row=x1,column=8).value)
    CR.append(sheet_obj.cell(row=x1,column=9).value)
    QR.append(sheet_obj.cell(row=x1,column=10).value)
    DER.append(sheet_obj.cell(row=x1,column=11).value)
    CP.append(sheet_obj.cell(row=x1,column=12).value)
    MCap.append(sheet_obj.cell(row=x1,column=13).value)
    Total_Returns.append(sheet_obj.cell(row=x1,column=14).value)
    EPS.append(sheet_obj.cell(row=x1, column=15).value)
    date2.append(sheet_obj.cell(row=x1,column=2).value)
    x1+=1
    x+=1
    z+= 1
    if z==7:
        z=0
        x+=3
        x1+=3
        y+=1
    if y>=48:
        break
print RD
print current_assets
print z
wb=openpyxl.Workbook()
ws1 = wb.create_sheet("Sheet_A")
ws1.title = "Process"
for i in range(1,len(RD)):
    c1=ws1.cell(row=i,column=1)
    c1.value=Company[i]
    c2 = ws1.cell(row=i, column=2)
    c2.value=date[i]
    c3=ws1.cell(row=i,column=3)
    c3.value=RD[i]
    c5 = ws1.cell(row=i, column=5)
    c5.value = date2[i]
    c6 = ws1.cell(row=i, column=6)
    c6.value = current_assets[i]
    c7 = ws1.cell(row=i, column=7)
    c7.value = current_liabilities[i]
    c8 = ws1.cell(row=i, column=8)
    c8.value = ROA[i]
    c9 = ws1.cell(row=i, column=9)
    c9.value = ROE[i]
    c10 = ws1.cell(row=i, column=10)
    c10.value = ROC[i]
    c11 = ws1.cell(row=i, column=11)
    c11.value = CR[i]
    c12 = ws1.cell(row=i, column=12)
    c12.value = QR[i]
    c13 = ws1.cell(row=i, column=13)
    c13.value = DER[i]
    c14 = ws1.cell(row=i, column=14)
    c14.value = CP[i]
    c15 = ws1.cell(row=i, column=15)
    c15.value = MCap[i]
    c16 = ws1.cell(row=i, column=16)
    c16.value = Total_Returns[i]
    c17 = ws1.cell(row=i, column=17)
    c17.value = EPS[i]
wb.template=False
wb.save(filename="sample.xlsx")
#wb.save(filename="Pharma_2009-2019_Dataset_Pooled.xlsx")
#wb_obj.save(filename = 'sample_book.xlsx')

#for i in range(2,38):
    #Return.append(((sheet_obj.cell(row = i, column = 1).value)))
    #RD.append(((sheet_obj.cell(row = i, column = 4).value)))
##Return = np.array(Return)
#D=np.array(RD)
#print kurtosis(Return)
#print skew(Return)
#print np.corrcoef(Return, RD)
