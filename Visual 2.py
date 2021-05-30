import openpyxl
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import scipy.stats as stats
from scipy.stats import spearmanr
path="/Users/yashjain/Desktop/Pharma_Dataset (1).xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
Return=[]
RD_TO_EBIT=[]
for i in range(2,62):
    Return.append(((sheet_obj.cell(row = i, column = 4).value)))
    RD_TO_EBIT.append(((sheet_obj.cell(row = i, column = 8).value)))
#print RD_TO_EBIT
Return = np.array(Return)
RD_TO_EBIT=np.array(RD_TO_EBIT)
mu=Return.mean()
std= Return.std()
x=np.linspace(mu-3*std,mu+3*std,100)
plt.plot(x,stats.norm.pdf(x,mu,std))
y=np.linspace(-1.0,2,10)
x=(y*0)+mu
plt.plot(x,y,color='red')
x2=np.linspace(-1.5,2,10)
y2=RD_TO_EBIT.mean()
y2=(x2*0)+y2
plt.plot(x2,y2,color='red')
plt.scatter(Return,RD_TO_EBIT,marker="*")
plt.xlabel("Annual Return")
plt.ylabel("%of EBIT RE-Invested in R&D")
plt.tight_layout()
plt.grid(True)
plt.legend(loc=1)
plt.show()
print np.corrcoef(Return, RD_TO_EBIT)



