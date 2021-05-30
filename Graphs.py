import openpyxl
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
path="/Users/yashjain/Desktop/Pharma_Dataset (1).xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
print(sheet_obj.max_column)
date=[]
RETURN_ANNUAL=[]
r=[]
R_TO_EBIT=[]
for i in range(2,7):
    date.append(((sheet_obj.cell(row = i, column = 3).value)))
    RETURN_ANNUAL.append((sheet_obj.cell(row = i, column = 4).value))
    r.append((sheet_obj.cell(row = i, column = 5).value))
    R_TO_EBIT.append((sheet_obj.cell(row = i, column = 8).value))
print r
reg=np.polyfit(x,y,deg=7)
ry=np.polyval(reg,x)

#plt.xticks(np.arange(np.datetime64('2014'), np.datetime64('2020')))
plt.plot(date,RETURN_ANNUAL,marker='*',lw=1.9,color='red')
#plt.plot(date,r,marker='.',lw=1.9,color='blue')
#plt.plot(date,R_TO_EBIT,lw=1.9,color='green')
plt.grid(True)
plt.legend(loc='upper left', frameon=False)
plt.show()